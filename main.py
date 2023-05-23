import pandas as pd
import openpyxl, json, os

variable_number_dict = {

    "Operator": 3,

    "Environment Temp": 9,

    "Barometer": 1,

    "Influent Temp": 141,
    "Influent PH": 137,
    "Influent D.O.": 113,
    "Influent Comp P.H.": 135,
    "Influent ALK": 101,
    "Influent SS": 139,

    "Final Effluent Temp": 2277,
    "Final Effluent PH": 2207,
    "Final Effluent D.O.": 2223,
    "Final Effluent Comp P.H.": 2205,
    "Final Effluent ALK": 2281,
    "Final Effluent SS": 2227,

    "#2 D Box Temp": 921,
    "#2 D Box PH": 917,
    "#2 D Box D.O.": 915,
    "#2 D Box 5 Mins": 925,
    "#2 D Box 10 Mins": 927,
    "#2 D Box 15 Mins": 929,
    "#2 D Box 20 Mins": 931,
    "#2 D Box 25 Mins": 933,
    "#2 D Box 30 Mins": 935,
    "#2 D Box 60 Mins": 937,

    "T.S.S Influent": 151,
    "T.S.S Primary": 729,
    "T.S.S Secondary": 1807,
    "T.S.S Effluent.": 2225,
    "T.S.S Baker": 2705,

    "T.S.S MLSS": 901,
    "T.S.S MLVSS": 905,
    "T.S.S RASS": 1707,
    "T.S.S RASVSS": 1711,
    "T.S.S SVI": "TSS SVI",

    "Reaction Tanks Sec P/H. Grab": 3101,
    "Reaction Tanks Sec Meter": 3103,
    "Reaction Tanks RT #1 PH": 3105,
    "Reaction Tanks TUBA (METER)": 3107,
    "Reaction Tanks RT #4 PH": 3109,

    "Reaction Tanks RAS PH": 1725,

    "COD Influent": 109,
    "COD Baker": 2703,

    "Ecoli": 2279,

    "BOD Influent Influent": 105,
    "BOD Primary": 703,
    "BOD Secondary": 1803,
    "BOD Effluent": 2209,
    "BOD Baker": 2701,

    "Comp. P.H./ALK Primary P.H.": 721,
    "Comp. P.H./ALK Primary ALK": 701,
    "Comp. P.H./ALK Secondary P.H.": 1829,
    "Comp. P.H./ALK Secondary ALK": 1831,

    "Baker Comp. P.H.": 2707,
    "Baker Grab P.H.": 2709,

    "Total P Secondary": 1817,
    "Total P Final Effluent": 2245,

    "Aluminum Final Effluent": 2291,
    "Aluminum H20 Department": 2509,
    "Copper Final Effluent": 2285,

    "Comag Influent Flow": 2001,
    "Comag Wasting Flow": 2003,

    "WAS Flow Set Point": 1705,
    "WAS Flow Actual": 1703,

    "Primary Sludge Q": 607,
    "Ras Q": 1701,

    "Ammonia Final Effluent": 2261,
    "Ammonia Nitrite": 2271,
    "Ammonia Nitrate": 2273,

    "Comag WAS PH": 2005,

    "Chlorine Residuals HI CHLORINE RESIDUAL mg/l": 2101,
    "Chlorine Residuals FINAL EFF.  CHLORINE ug/L": 2275,

    "Fournier Press 1 Solids %": 2413,
    "Fournier Press 2 Solids %": 2415,
    "Fournier Press Flow": 2405,

    "PPMV": 2011,

    "Turbidity Meter": 2283,
    "Turbidity Secondary": 1841,
    "Turbidity Final Effluent": 2007,
    "Turbidity Comag": 2009,
    "Turbidity Baker Composit": 2713,
    "Turbidity Baker Grab": 2711,

    "Primary Clarifier 1 Depth of Blanket": 305,
    "Primary Clarifier 2 Depth of Blanket": 405,
    "Primary Clarifier 3 Depth of Blanket": 505,

    "Secondary Clarifier 1 Depth of Blanket": 1205,
    "Secondary Clarifier 2 Depth of Blanket": 1305,
    "Secondary Clarifier 3 Depth of Blanket": 1405,
    "Secondary Clarifier 4 Depth of Blanket": 1505,

    "Tertiary Clarifier 1 Depth of Blanket": 2013,
    "Tertiary Clarifier 2 Depth of Blanket": 2017,

    "Gravity Thickener 1 Depth of Blanket": 2401,
    "Gravity Thickener 2 Depth of Blanket": 2403,

    "Flow Average": 115,
    "Flow Max": 117,
    "Flow Min": 119,

    "Plant Chemicals 7500 CAUSTIC Tank Level Today": 2307,
    "Plant Chemicals 6000 HYPO Tank Level Today": 2313,
    "Plant Chemicals 5000 BISULFITE Tank Level Today": 2325,
    "Plant Chemicals 5000 SODIUM ALUM Tank Level Today": 2319,
    "Plant Chemicals PRESS POLYMER Tank Level Today": 2359,

    "Comag 3500 CAUSTIC Tank Level Today": 2351,
    "Comag 7500 ALUM Tank Level Today": 2301,
    "Comag  POLYMER Tank Level Today": 2335,

    "Error": "Error - A value was pulled from a spreadsheet that does not exist in the variable_number_dict",
    "None": "Error - A value was pulled from a spreadsheet that does not exist in the variable_number_dict"

}

total_data_dict = {'Date': [], 'Day': [], 'Operator': [], 'Temp': [], 'Barometer': [], 'Influent Time': [],
                   'Influent Temp': [],
                   'Influent PH': [], 'Influent D.O.': [], 'Influent Comp P.H.': [], 'Influent ALK': [],
                   'Influent SS': [],
                   'Final Effluent Time': [], 'Final Effluent Temp': [], 'Final Effluent PH': [],
                   'Final Effluent D.O.': [],
                   'Final Effluent Comp P.H.': [], 'Final Effluent ALK': [], 'Final Effluent SS': [],
                   '#2 D Box Time': [],
                   '#2 D Box Temp': [], '#2 D Box PH': [], '#2 D Box D.O.': [], '#2 D Box 5 Mins': [],
                   '#2 D Box 10 Mins': [],
                   '#2 D Box 15 Mins': [], '#2 D Box 20 Mins': [], '#2 D Box 25 Mins': [], '#2 D Box 30 Mins': [],
                   '#2 D Box 60 Mins': [], 'T.S.S Influent': [], 'T.S.S Primary': [], 'T.S.S Secondary': [],
                   'T.S.S Effluent.': [],
                   'T.S.S Baker': [], 'T.S.S % Removal': [], 'T.S.S MLSS': [], 'T.S.S MLVSS': [], 'T.S.S RASS': [],
                   'T.S.S RASVSS': [], 'T.S.S SVI': [], 'Reaction Tanks Sec P/H. Grab': [],
                   'Reaction Tanks Sec Meter': [],
                   'Reaction Tanks RT #1 PH': [], 'Reaction Tanks TUBA (METER)': [], 'Reaction Tanks RT #4 PH': [],
                   'Reaction Tanks RAS PH': [], 'COD Influent': [], 'COD Baker': [], 'Ecoli': [],
                   'BOD Influent Influent': [],
                   'BOD Influent Primary': [], 'BOD Influent Secondary': [], 'BOD Influent Effluent': [],
                   'BOD Influent Baker': [],
                   'Comp. P.H./ALK Primary P.H.': [], 'Comp. P.H./ALK Primary ALK': [],
                   'Comp. P.H./ALK Secondary P.H.': [],
                   'Comp. P.H./ALK Secondary ALK': [], 'Baker Comp. P.H.': [], 'Baker Grab P.H.': [],
                   'Total P Secondary': [],
                   'Total P Final Effluent': [], 'Aluminum Final Effluent': [], 'Aluminum H20 Department': [],
                   'Comag Influent Flow': [], 'Comag Wasting Flow': [], 'WAS Flow Set Point': [], 'WAS Flow Actual': [],
                   'Primary Sludge Q': [], 'Ras Q': [], 'Ammonia Final Effluent': [], 'Ammonia Nitrite': [],
                   'Ammonia Nitrate': [],
                   'Comag WAS PH': [], 'Chlorine Residuals Time': [], 'Chlorine Residuals Operator': [],
                   'Chlorine Residuals Q EFFLUENT CHEMICAL CONTROL': [],
                   'Chlorine Residuals HYPO PUMP ONLINE A OR H': [],
                   'Chlorine Residuals HYPO TANK LEVEL': [], 'Chlorine Residuals RESIDUAL SHED 1': [],
                   'Chlorine Residuals RESIDUAL SHED 2': [], 'Chlorine Residuals RESIDUAL SHED 3': [],
                   'Chlorine Residuals BISULFITE PUMP ONLINE A OR H': [], 'Chlorine Residuals BISULFITE TANK LEVEL': [],
                   'Chlorine Residuals P.H.': [], 'Chlorine Residuals HI CHLORINE RESIDUAL mg/l': [],
                   'Chlorine Residuals FINAL EFF.  CHLORINE ug/L': [], 'Fournier Press 1 Solids %': [],
                   'Fournier Press 2 Solids %': [], 'Fournier Press Flow': [], 'PPMV': [], 'Turbidity Meter': [],
                   'Turbidity Secondary': [], 'Turbidity Final Effluent': [], 'Turbidity Comag': [],
                   'Turbidity Baker Composit': [], 'Turbidity Baker Grab': [],
                   'Primary Clarifier 1 Depth of Blanket': [],
                   'Primary Clarifier 2 Depth of Blanket': [], 'Primary Clarifier 3 Depth of Blanket': [],
                   'Secondary Clarifier 1 Depth of Blanket': [], 'Secondary Clarifier 2 Depth of Blanket': [],
                   'Secondary Clarifier 3 Depth of Blanket': [], 'Secondary Clarifier 4 Depth of Blanket': [],
                   'Tertiary Clarifier 1 Depth of Blanket': [], 'Tertiary Clarifier 2 Depth of Blanket': [],
                   'Gravity Thickener 1 Depth of Blanket': [], 'Gravity Thickener 2 Depth of Blanket': [],
                   'Flow Average': [],
                   'Flow Max': [], 'Flow Min': [], 'Ras Pump 1': [], 'Ras Pump 2': [], 'Ras Pump 3': [],
                   'Ras Pump 4': [],
                   'Ras Pump 5': [], 'Plant Chemicals 7500 CAUSTIC Tank Level Today': [],
                   'Plant Chemicals 6000 HYPO Tank Level Today': [],
                   'Plant Chemicals 5000 BISULFITE Tank Level Today': [],
                   'Plant Chemicals 5000 SODIUM ALUM Tank Level Today': [],
                   'Plant Chemicals PRESS POLYMER Tank Level Today': [],
                   'Plant Chemicals 7500 CAUSTIC Tank Level Yesterday': [],
                   'Plant Chemicals 6000 HYPO Tank Level Yesterday': [],
                   'Plant Chemicals 5000 BISULFITE Tank Level Yesterday': [],
                   'Plant Chemicals 5000 SODIUM ALUM Tank Level Yesterday': [],
                   'Plant Chemicals PRESS POLYMER Tank Level Yesterday': [],
                   'Plant Chemicals 7500 CAUSTIC Tank Level User': [],
                   'Plant Chemicals 6000 HYPO Tank Level User': [],
                   'Plant Chemicals 5000 BISULFITE Tank Level User': [],
                   'Plant Chemicals 5000 SODIUM ALUM Tank Level User': [],
                   'Plant Chemicals PRESS POLYMER Tank Level User': [],
                   'Comag 3500 CAUSTIC Tank Level Today': [], 'Comag 7500 ALUM Tank Level Today': [],
                   'Comag  POLYMER Tank Level Today': [], 'Comag 3500 CAUSTIC Tank Level Yesterday': [],
                   'Comag 7500 ALUM Tank Level Yesterday': [], 'Comag  POLYMER Tank Level Yesterday': [],
                   'Comag 3500 CAUSTIC Tank Level Used': [], 'Comag 7500 ALUM Tank Level Used': [],
                   'Comag  POLYMER Tank Level Used': []}


def read_Excel_File(ws):
    return_dict = {
        "Date": str(ws["C5"].value)[:10].replace(".", "-"),

        "Day": str(ws["E5"].value),

        "Operator": str(ws["G5"].value),

        "Temp": str(ws["I5"].value),

        "Barometer": str(ws["K5"].value),

        "Influent Time": str(ws["C7"].value),
        "Influent Temp": str(ws["C8"].value),
        "Influent PH": str(ws["C9"].value),
        "Influent D.O.": str(ws["C10"].value),
        "Influent Comp P.H.": str(ws["C11"].value),
        "Influent ALK": str(ws["C12"].value),
        "Influent SS": str(ws["C13"].value),

        "Final Effluent Time": str(ws["E7"].value),
        "Final Effluent Temp": str(ws["E8"].value),
        "Final Effluent PH": str(ws["E9"].value),
        "Final Effluent D.O.": str(ws["E10"].value),
        "Final Effluent Comp P.H.": str(ws["E11"].value),
        "Final Effluent ALK": str(ws["E12"].value),
        "Final Effluent SS": str(ws["E13"].value),

        "#2 D Box Time": str(ws["C15"].value),
        "#2 D Box Temp": str(ws["C16"].value),
        "#2 D Box PH": str(ws["C17"].value),
        "#2 D Box D.O.": str(ws["C18"].value),
        "#2 D Box 5 Mins": str(ws["C19"].value),
        "#2 D Box 10 Mins": str(ws["C20"].value),
        "#2 D Box 15 Mins": str(ws["C21"].value),
        "#2 D Box 20 Mins": str(ws["C22"].value),
        "#2 D Box 25 Mins": str(ws["C23"].value),
        "#2 D Box 30 Mins": str(ws["C24"].value),
        "#2 D Box 60 Mins": str(ws["C25"].value),

        "T.S.S Influent": str(ws["E15"].value),
        "T.S.S Primary": str(ws["E16"].value),
        "T.S.S Secondary": str(ws["E17"].value),
        "T.S.S Effluent.": str(ws["E18"].value),
        "T.S.S Baker": str(ws["E19"].value),
        "T.S.S % Removal": str(ws["E20"].value),
        "T.S.S MLSS": str(ws["E21"].value),
        "T.S.S MLVSS": str(ws["E22"].value),
        "T.S.S RASS": str(ws["E23"].value),
        "T.S.S RASVSS": str(ws["E24"].value),
        "T.S.S SVI": str(ws["E25"].value),

        "Reaction Tanks Sec P/H. Grab": str(ws["G7"].value),
        "Reaction Tanks Sec Meter": str(ws["G8"].value),
        "Reaction Tanks RT #1 PH": str(ws["G9"].value),
        "Reaction Tanks TUBA (METER)": str(ws["G10"].value),
        "Reaction Tanks RT #4 PH": str(ws["G11"].value),

        "Reaction Tanks RAS PH": str(ws["G13"].value),

        "COD Influent": str(ws["G15"].value),
        "COD Baker": str(ws["G16"].value),

        "Ecoli": str(ws["G17"].value),

        "BOD Influent Influent": str(ws["G20"].value),
        "BOD Influent Primary": str(ws["G21"].value),
        "BOD Influent Secondary": str(ws["G22"].value),
        "BOD Influent Effluent": str(ws["G23"].value),
        "BOD Influent Baker": str(ws["G24"].value),

        "Comp. P.H./ALK Primary P.H.": str(ws["I7"].value),
        "Comp. P.H./ALK Primary ALK": str(ws["I8"].value),
        "Comp. P.H./ALK Secondary P.H.": str(ws["I9"].value),
        "Comp. P.H./ALK Secondary ALK": str(ws["I10"].value),

        "Baker Comp. P.H.": str(ws["H13"].value),
        "Baker Grab P.H.": str(ws["I13"].value),

        "Total P Secondary": str(ws["I15"].value),
        "Total P Final Effluent": str(ws["I15"].value),

        "Aluminum Final Effluent": str(ws["I19"].value),
        "Aluminum H20 Department": str(ws["I20"].value),

        "Comag Influent Flow": str(ws["H23"].value),
        "Comag Wasting Flow": str(ws["H24"].value),

        "WAS Flow Set Point": str(ws["K7"].value),
        "WAS Flow Actual": str(ws["K8"].value),

        "Primary Sludge Q": str(ws["J10"].value),
        "Ras Q": str(ws["J12"].value),

        "Ammonia Final Effluent": str(ws["K15"].value),
        "Ammonia Nitrite": str(ws["K16"].value),
        "Ammonia Nitrate": str(ws["K17"].value),

        "Comag WAS PH": str(ws["J20"].value),

        "Chlorine Residuals Time": str(ws["M6"].value),
        "Chlorine Residuals Operator": str(ws["O6"].value),
        "Chlorine Residuals Q EFFLUENT CHEMICAL CONTROL": str(ws["O7"].value),
        "Chlorine Residuals HYPO PUMP ONLINE A OR H": str(ws["O8"].value),
        "Chlorine Residuals HYPO TANK LEVEL": str(ws["O9"].value),
        "Chlorine Residuals RESIDUAL SHED 1": str(ws["O10"].value),
        "Chlorine Residuals RESIDUAL SHED 2": str(ws["O11"].value),
        "Chlorine Residuals RESIDUAL SHED 3": str(ws["O12"].value),
        "Chlorine Residuals BISULFITE PUMP ONLINE A OR H": str(ws["O13"].value),
        "Chlorine Residuals BISULFITE TANK LEVEL": str(ws["O14"].value),
        "Chlorine Residuals P.H.": str(ws["O15"].value),
        "Chlorine Residuals HI CHLORINE RESIDUAL mg/l": str(ws["O16"].value),
        "Chlorine Residuals FINAL EFF.  CHLORINE ug/L": str(ws["O17"].value),

        # Needs to split or regex or something.
        # "Fournier Press 1 Feed %": str(ws["L20"].value)[],
        # "Fournier Press 2 Feed %": str(ws["L20"].value).rstrip('/'),

        "Fournier Press 1 Solids %": str(ws["M20"].value),
        "Fournier Press 2 Solids %": str(ws["N20"].value),
        "Fournier Press Flow": str(ws["O20"].value),

        "PPMV": str(ws["P24"].value),

        "Turbidity Meter": str(ws["P27"].value),
        "Turbidity Secondary": str(ws["P28"].value),
        "Turbidity Final Effluent": str(ws["P29"].value),
        "Turbidity Comag": str(ws["P30"].value),
        "Turbidity Baker Composit": str(ws["P31"].value),
        "Turbidity Baker Grab": str(ws["P32"].value),

        "Primary Clarifier 1 Depth of Blanket": str(ws["L27"].value),
        "Primary Clarifier 2 Depth of Blanket": str(ws["L28"].value),
        "Primary Clarifier 3 Depth of Blanket": str(ws["L29"].value),

        "Secondary Clarifier 1 Depth of Blanket": str(ws["J27"].value),
        "Secondary Clarifier 2 Depth of Blanket": str(ws["J28"].value),
        "Secondary Clarifier 3 Depth of Blanket": str(ws["J29"].value),
        "Secondary Clarifier 4 Depth of Blanket": str(ws["J30"].value),

        "Tertiary Clarifier 1 Depth of Blanket": str(ws["N27"].value),
        "Tertiary Clarifier 2 Depth of Blanket": str(ws["N28"].value),

        "Gravity Thickener 1 Depth of Blanket": str(ws["N31"].value),
        "Gravity Thickener 2 Depth of Blanket": str(ws["N32"].value),

        "Flow Average": str(ws["H32"].value),
        "Flow Max": str(ws["J32"].value),
        "Flow Min": str(ws["L32"].value),

        "Ras Pump 1": str(ws["H27"].value),
        "Ras Pump 2": str(ws["H28"].value),
        "Ras Pump 3": str(ws["H29"].value),
        "Ras Pump 4": str(ws["H30"].value),
        "Ras Pump 5": str(ws["H31"].value),

        "Plant Chemicals 7500 CAUSTIC Tank Level Today": str(ws["D28"].value),
        "Plant Chemicals 6000 HYPO Tank Level Today": str(ws["D29"].value),
        "Plant Chemicals 5000 BISULFITE Tank Level Today": str(ws["D30"].value),
        "Plant Chemicals 5000 SODIUM ALUM Tank Level Today": str(ws["D31"].value),
        "Plant Chemicals PRESS POLYMER Tank Level Today": str(ws["D32"].value),

        "Plant Chemicals 7500 CAUSTIC Tank Level Yesterday": str(ws["E28"].value),
        "Plant Chemicals 6000 HYPO Tank Level Yesterday": str(ws["E29"].value),
        "Plant Chemicals 5000 BISULFITE Tank Level Yesterday": str(ws["E30"].value),
        "Plant Chemicals 5000 SODIUM ALUM Tank Level Yesterday": str(ws["E31"].value),
        "Plant Chemicals PRESS POLYMER Tank Level Yesterday": str(ws["E32"].value),

        "Plant Chemicals 7500 CAUSTIC Tank Level User": str(ws["F28"].value),
        "Plant Chemicals 6000 HYPO Tank Level User": str(ws["F29"].value),
        "Plant Chemicals 5000 BISULFITE Tank Level User": str(ws["F30"].value),
        "Plant Chemicals 5000 SODIUM ALUM Tank Level User": str(ws["F31"].value),
        "Plant Chemicals PRESS POLYMER Tank Level User": str(ws["F32"].value),

        "Comag 3500 CAUSTIC Tank Level Today": str(ws["L23"].value),
        "Comag 7500 ALUM Tank Level Today": str(ws["L24"].value),
        "Comag  POLYMER Tank Level Today": str(ws["L25"].value),

        "Comag 3500 CAUSTIC Tank Level Yesterday": str(ws["M23"].value),
        "Comag 7500 ALUM Tank Level Yesterday": str(ws["M24"].value),
        "Comag  POLYMER Tank Level Yesterday": str(ws["M25"].value),

        "Comag 3500 CAUSTIC Tank Level Used": str(ws["N23"].value),
        "Comag 7500 ALUM Tank Level Used": str(ws["N24"].value),
        "Comag  POLYMER Tank Level Used": str(ws["N25"].value),

    }

    return return_dict

if __name__ == '__main__':

    data_list = []

    spreadsheet_tab_to_ignore = ["BLANK LAB SHEET", "F M", "Detention times", "Fournier Press", "SRT Time",
                                 "DMR Weekly",
                                 "Sheet 1",
                                 "DMR"]

    spreadsheets_to_iterate_through = os.listdir("Daily Lab Sheets")

    for spreadsheet in spreadsheets_to_iterate_through:
        wb = openpyxl.load_workbook("Daily Lab Sheets/" + spreadsheet)

        workbook_sheets = [sheet.title for sheet in wb.worksheets if sheet.title not in spreadsheet_tab_to_ignore]



        for sheet in workbook_sheets:
            print(sheet)
            print(wb)
            work_sheet = wb[sheet]

            the_date = str(work_sheet["C5"].value)

            if the_date is None:
                print("Skipping " + sheet + " because it does not have a value in C5")
                continue
            print("Getting the data for: " + sheet)

            tab_data = read_Excel_File(work_sheet)



            for account_name, account_value in tab_data.items():

                try:
                    account_variable_number = variable_number_dict[account_name]
                except KeyError:
                    account_variable_number = "N/A"

                data_list.append((account_variable_number, the_date, account_value, account_name))

    print(data_list)

    df = pd.DataFrame(data_list, columns=("Date","Variable Number", "Value", "Account Name"))
    df.to_csv("Export.csv")

